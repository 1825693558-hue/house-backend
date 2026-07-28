"""
房源导出服务 - 异步生成 Excel + 图片/视频 ZIP 压缩包

流程:
  1. 查询全部房源（含联系人、家电、小区信息）
  2. 生成 Excel 文件
  3. 逐个房源下载图片/视频到对应文件夹
  4. 打包为 ZIP
  5. 前端轮询任务状态，完成后下载
"""
import io
import os
import re
import uuid
import zipfile
import threading
import time
import logging
from datetime import datetime
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy.orm import Session, selectinload

from app.core.security import aes_decrypt
from app.models.house import House
from app.models.community import Community
from app.models.contact import Contact
from app.models.house_appliance import HouseAppliance
from app.models.appliance import Appliance

logger = logging.getLogger(__name__)

# ---------- 任务状态管理 ----------

class ExportTask:
    """单个导出任务的状态"""

    def __init__(self, task_id: str, export_type: str, filters: dict | None = None):
        self.task_id = task_id
        self.export_type = export_type  # "all" | "filtered"
        self.filters = filters or {}
        self.status = "pending"        # pending → processing → done / failed
        self.progress = 0              # 0-100
        self.message = "等待开始..."
        self.zip_path: str | None = None
        self.zip_filename: str | None = None
        self.total_houses = 0
        self.processed_houses = 0
        self.created_at = datetime.now()
        self.error: str | None = None

    def to_dict(self) -> dict:
        return {
            "task_id": self.task_id,
            "status": self.status,
            "progress": self.progress,
            "message": self.message,
            "total_houses": self.total_houses,
            "processed_houses": self.processed_houses,
            "created_at": self.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            "error": self.error,
        }


# 全局任务存储（内存中，进程重启后丢失）
_tasks: dict[str, ExportTask] = {}
_tasks_lock = threading.Lock()

# 临时文件目录
EXPORT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "exports")
os.makedirs(EXPORT_DIR, exist_ok=True)


def create_task(export_type: str, filters: dict | None = None) -> str:
    """创建导出任务，返回 task_id"""
    task_id = uuid.uuid4().hex[:12]
    task = ExportTask(task_id, export_type, filters)
    with _tasks_lock:
        _tasks[task_id] = task
    return task_id


def get_task(task_id: str) -> ExportTask | None:
    with _tasks_lock:
        return _tasks.get(task_id)


def _cleanup_old_tasks():
    """清理超过 2 小时的已完成任务及其文件"""
    cutoff = datetime.now().timestamp() - 7200
    with _tasks_lock:
        to_remove = []
        for tid, task in _tasks.items():
            if task.status in ("done", "failed") and task.created_at.timestamp() < cutoff:
                if task.zip_path and os.path.exists(task.zip_path):
                    try:
                        os.remove(task.zip_path)
                    except OSError:
                        pass
                to_remove.append(tid)
        for tid in to_remove:
            del _tasks[tid]


# ---------- 文件名安全处理 ----------

def _sanitize_folder_name(name: str) -> str:
    """将文件夹名中的非法字符替换为下划线"""
    return re.sub(r'[<>:"/\\|?*]', '_', name).strip().rstrip('.')


def _build_house_folder_name(house: House, community_name: str | None, used_names: set) -> str:
    """构建房源文件夹名：【小区·地址】，重名时追加 ID"""
    parts = []
    if community_name:
        parts.append(community_name)
    if house.address:
        parts.append(house.address)

    if not parts:
        folder = f"房源_{house.id}"
    else:
        folder = _sanitize_folder_name("·".join(parts))

    base = f"【{folder}】"
    if base in used_names:
        base = f"【{folder}·ID{house.id}】"
    used_names.add(base)
    return base


# ---------- Excel 生成 ----------

# 表头样式
_header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
_header_fill = PatternFill(start_color="2d8f5e", end_color="2d8f5e", fill_type="solid")
_header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
_cell_align = Alignment(vertical="center", wrap_text=True)
_thin_border = Border(
    left=Side(style="thin", color="d0d0d0"),
    right=Side(style="thin", color="d0d0d0"),
    top=Side(style="thin", color="d0d0d0"),
    bottom=Side(style="thin", color="d0d0d0"),
)

# Excel 列定义: (列名, 取值函数)
EXCEL_COLUMNS: list[tuple[str, Any]] = [
    ("编号",            lambda h, ctx: h["id"]),
    ("小区名称",        lambda h, ctx: h["community_name"] or ""),
    ("小区地址",        lambda h, ctx: h["community_address"] or ""),
    ("楼号/门牌",       lambda h, ctx: h["address"] or ""),
    ("面积(㎡)",        lambda h, ctx: h["area"] or ""),
    ("所在楼层",        lambda h, ctx: h["floor"] or ""),
    ("总楼层",          lambda h, ctx: h["total_floors"] or ""),
    ("出售价(万)",      lambda h, ctx: h["sale_price"] or ""),
    ("出租价(元/月)",   lambda h, ctx: h["rent_price"] or ""),
    ("价格备注",        lambda h, ctx: h["price_note"] or ""),
    ("房源状态",        lambda h, ctx: h["status"] or ""),
    ("房源类型",        lambda h, ctx: h["house_type"] or ""),
    ("装修状况",        lambda h, ctx: h["decoration"] or ""),
    ("钥匙类型",        lambda h, ctx: h["key_type"] or ""),
    ("密码锁密码",      lambda h, ctx: h["lock_password"] or ""),
    ("家电配置",        lambda h, ctx: ", ".join(h["appliance_names"])),
    ("联系人",          lambda h, ctx: "\n".join(h["contact_lines"])),
    ("主联系人",        lambda h, ctx: h["primary_contact"] or ""),
    ("房源描述",        lambda h, ctx: h["description"] or ""),
    ("图片数量",        lambda h, ctx: len(h["images"] or [])),
    ("是否有视频",      lambda h, ctx: "是" if h["video_url"] else "否"),
    ("文件夹名称",      lambda h, ctx: h["folder_name"]),
    ("创建时间",        lambda h, ctx: h["created_at"]),
    ("更新时间",        lambda h, ctx: h["updated_at"]),
]


def _generate_excel(house_data_list: list[dict]) -> bytes:
    """生成 Excel 文件，返回字节内容"""
    wb = Workbook()
    ws = wb.active
    ws.title = "房源信息"

    # 写表头
    headers = [col[0] for col in EXCEL_COLUMNS]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = _header_align
        cell.border = _thin_border

    # 写数据行
    for house_data in house_data_list:
        row = []
        for _, getter in EXCEL_COLUMNS:
            row.append(getter(house_data, {}))
        ws.append(row)

    # 设置数据行样式
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = _cell_align
            cell.border = _thin_border

    # 列宽
    col_widths = [8, 16, 24, 20, 10, 10, 10, 12, 14, 16, 10, 10, 10, 12, 14, 24, 30, 16, 40, 10, 10, 30, 20, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = w

    # 冻结首行
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------- 媒体文件下载 ----------

def _download_media(url: str, timeout: float = 30.0) -> bytes | None:
    """下载媒体文件，返回字节内容，失败返回 None"""
    if not url:
        return None
    try:
        resp = requests.get(url, timeout=timeout, allow_redirects=True)
        if resp.status_code == 200:
            return resp.content
        logger.warning("下载媒体失败 %s: HTTP %d", url, resp.status_code)
    except Exception as e:
        logger.warning("下载媒体异常 %s: %s", url, e)
    return None


def _get_file_extension(url: str, default: str = ".jpg") -> str:
    """从 URL 提取文件扩展名"""
    # 去掉查询参数
    path = url.split("?")[0]
    _, ext = os.path.splitext(path)
    if ext and len(ext) <= 6:
        return ext.lower()
    return default


# ---------- 核心导出逻辑 ----------

def _query_all_houses(db: Session, filters: dict | None = None) -> list[House]:
    """查询全部房源（含关联数据），支持筛选"""
    query = (
        db.query(House)
        .options(
            selectinload(House.community),
            selectinload(House.contacts),
            selectinload(House.house_appliances).selectinload(HouseAppliance.appliance),
        )
        .filter(House.is_deleted == False)  # noqa: E712
    )

    filters = filters or {}

    # 复用列表接口的筛选逻辑
    if filters.get("keyword"):
        kw = f"%{filters['keyword']}%"
        query = query.join(Community, House.community_id == Community.id, isouter=True).filter(
            (Community.name.like(kw)) | (House.address.like(kw))
        )

    if filters.get("status"):
        statuses = [s.strip() for s in filters["status"].split(",") if s.strip()]
        if statuses:
            query = query.filter(House.status.in_(statuses))

    if filters.get("decoration"):
        query = query.filter(House.decoration == filters["decoration"])

    if filters.get("key_type"):
        query = query.filter(House.key_type == filters["key_type"])

    if filters.get("community_id"):
        query = query.filter(House.community_id == filters["community_id"])

    if filters.get("house_use_type") == "sale":
        query = query.filter(House.sale_price.isnot(None))
    elif filters.get("house_use_type") == "rent":
        query = query.filter(House.rent_price.isnot(None))

    return query.order_by(House.created_at.desc()).all()


def _build_house_data(house: House, folder_name: str) -> dict:
    """将 House ORM 对象转为导出用的字典"""
    # 小区信息
    community_name = house.community.name if house.community else None
    community_address = house.community.address if house.community else None

    # 联系人
    contact_lines = []
    primary_contact = ""
    for c in sorted(house.contacts, key=lambda x: (-x.is_primary, x.id)):
        phone_part = f": {c.phone}" if c.phone else ""
        line = f"{c.name}({c.role or '联系人'}){phone_part}"
        if c.is_primary:
            line += " ★"
            primary_contact = f"{c.name} {c.phone or ''}".strip()
        contact_lines.append(line)

    # 家电
    appliance_names = []
    for ha in house.house_appliances:
        name = ha.appliance.name if ha.appliance else f"家电#{ha.appliance_id}"
        if ha.note:
            name += f"({ha.note})"
        appliance_names.append(name)

    # 密码锁密码解密
    lock_password = ""
    if house.lock_password and house.key_type == "密码锁":
        try:
            lock_password = aes_decrypt(house.lock_password)
        except Exception:
            lock_password = "******"

    # 图片列表
    images = house.images or []
    if not images and house.media:
        images = [m.get("url") for m in house.media if m.get("type") == "image" and m.get("url")]

    return {
        "id": house.id,
        "community_name": community_name,
        "community_address": community_address,
        "address": house.address,
        "area": float(house.area) if house.area else None,
        "floor": house.floor,
        "total_floors": house.total_floors,
        "sale_price": float(house.sale_price) if house.sale_price else None,
        "rent_price": float(house.rent_price) if house.rent_price else None,
        "price_note": house.price_note,
        "status": house.status,
        "house_type": house.house_type,
        "decoration": house.decoration,
        "key_type": house.key_type,
        "lock_password": lock_password,
        "appliance_names": appliance_names,
        "contact_lines": contact_lines,
        "primary_contact": primary_contact,
        "description": house.description,
        "images": images,
        "video_url": house.video_url or "",
        "folder_name": folder_name,
        "created_at": house.created_at.strftime("%Y-%m-%d %H:%M") if house.created_at else "",
        "updated_at": house.updated_at.strftime("%Y-%m-%d %H:%M") if house.updated_at else "",
    }


def _run_export(task: ExportTask, db_session_factory):
    """实际执行导出逻辑（在后台线程中运行）"""
    task.status = "processing"
    task.message = "正在查询房源数据..."

    try:
        db = db_session_factory()

        # 1. 查询全部房源
        houses = _query_all_houses(db, task.filters)
        task.total_houses = len(houses)

        if not houses:
            task.status = "failed"
            task.error = "没有符合条件的房源数据"
            task.message = "导出失败：无数据"
            db.close()
            return

        task.message = f"共 {task.total_houses} 条房源，正在生成 Excel..."

        # 2. 构建导出数据
        used_folder_names: set = set()
        house_data_list = []
        for house in houses:
            folder_name = _build_house_folder_name(house, house.community.name if house.community else None, used_folder_names)
            house_data_list.append(_build_house_data(house, folder_name))

        # 3. 生成 Excel
        task.message = "正在生成 Excel 文件..."
        excel_bytes = _generate_excel(house_data_list)

        # 4. 创建 ZIP 并写入 Excel
        task.message = "正在打包，下载图片/视频..."
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        zip_filename = f"房源导出_{timestamp}.zip"
        zip_path = os.path.join(EXPORT_DIR, f"{task.task_id}_{zip_filename}")

        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            # 写入 Excel
            zf.writestr("房源信息.xlsx", excel_bytes)

            # 逐个房源下载图片/视频
            for i, hd in enumerate(house_data_list):
                folder = hd["folder_name"]
                images = hd["images"]
                video_url = hd["video_url"]

                # 计算需要下载的文件总数
                media_count = len(images) + (1 if video_url else 0)
                if media_count == 0:
                    task.processed_houses = i + 1
                    task.progress = int((i + 1) / task.total_houses * 100)
                    task.message = f"正在处理 {i + 1}/{task.total_houses}：{folder}"
                    continue

                task.message = f"正在下载 {i + 1}/{task.total_houses}：{folder}（{media_count}个文件）"

                # 下载图片
                for img_idx, img_url in enumerate(images):
                    img_data = _download_media(img_url)
                    if img_data:
                        ext = _get_file_extension(img_url, ".jpg")
                        img_name = f"{img_idx + 1}{ext}"
                        zf.writestr(f"{folder}/图片/{img_name}", img_data)

                # 下载视频
                if video_url:
                    video_data = _download_media(video_url, timeout=120.0)
                    if video_data:
                        ext = _get_file_extension(video_url, ".mp4")
                        zf.writestr(f"{folder}/视频/1{ext}", video_data)

                task.processed_houses = i + 1
                task.progress = int((i + 1) / task.total_houses * 100)

        # 5. 完成
        task.status = "done"
        task.progress = 100
        task.message = f"导出完成，共 {task.total_houses} 条房源"
        task.zip_path = zip_path
        task.zip_filename = zip_filename
        db.close()

        logger.info("导出任务 %s 完成: %s", task.task_id, zip_filename)

    except Exception as e:
        task.status = "failed"
        task.error = str(e)
        task.message = f"导出失败: {e}"
        logger.exception("导出任务 %s 失败", task.task_id)
    finally:
        _cleanup_old_tasks()


def start_export_task(task_id: str, db_session_factory):
    """在后台线程中启动导出任务"""
    task = get_task(task_id)
    if not task:
        return

    thread = threading.Thread(target=_run_export, args=(task, db_session_factory), daemon=True)
    thread.start()


def cleanup_task_file(task_id: str):
    """任务完成后清理文件（手动触发）"""
    task = get_task(task_id)
    if task and task.zip_path and os.path.exists(task.zip_path):
        try:
            os.remove(task.zip_path)
        except OSError:
            pass
